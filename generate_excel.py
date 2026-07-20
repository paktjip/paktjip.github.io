import os
import subprocess
import sys

# Auto-install openpyxl if not present
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl tidak ditemukan. Melakukan instalasi otomatis...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Generator Undangan WA"
    
    # Aktifkan grid lines agar sel terlihat rapi
    ws.views.sheetView[0].showGridLines = True
    
    # Base URL website undangan Anda (sesuaikan jika diubah ke domain custom)
    base_url = "https://paktjip.github.io/CobaUnd_Mono.html"
    
    # 1. Judul Halaman / Banner
    ws['A1'] = "GENERATOR PESAN & LINK UNDANGAN PERNIKAHAN LUKAS & KEZIA"
    ws['A1'].font = Font(name="Playfair Display", size=14, bold=True, color="3B5D3B")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A2'] = "Petunjuk: Aktifkan Apps Script (di Google Sheets) atau VBA Macro (di Excel) agar rumus otomatis berfungsi."
    ws['A2'].font = Font(name="Lato", size=9, italic=True, color="555555")
    ws.merge_cells('A2:E2')
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # 2. Header Tabel
    headers = ["No", "Nama Tamu (Kolom B)", "URL Undangan (Kolom C - Otomatis)", "No WhatsApp (Kolom D)", "Link Kirim WhatsApp (Kolom E - Otomatis)"]
    ws.append([]) # Baris kosong pembatas
    
    # Tambah headers
    ws.append(headers)
    
    # Format Headers (Baris 4)
    header_font = Font(name="Lato", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B5D3B", end_color="3B5D3B", fill_type="solid") # Sage Green
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_idx in range(1, 6):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # 3. Data Tamu Sampel
    sample_guests = [
        ("Budi Santoso", "628123456789"),
        ("Anita Wijaya", "628987654321"),
        ("Lukas & Partner", "628555555555"),
        ("Keluarga Bpk. Toni", "628111111111")
    ]
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Mulai menulis baris data dari baris 5
    for i, (name, wa_num) in enumerate(sample_guests, start=1):
        row_num = 4 + i
        
        # A: No
        ws.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal="center")
        
        # B: Nama Tamu
        ws.cell(row=row_num, column=2, value=name).alignment = Alignment(horizontal="left")
        
        # C: URL Undangan (RUMUS OTOMATIS)
        url_formula = f'="https://paktjip.github.io/CobaUnd_Mono.html?invite=" & ENCODE_BASE64(B{row_num})'
        ws.cell(row=row_num, column=3, value=url_formula).alignment = Alignment(horizontal="left")
        
        # D: No WhatsApp
        ws.cell(row=row_num, column=4, value=wa_num).alignment = Alignment(horizontal="center")
        
        # E: Link WhatsApp (RUMUS OTOMATIS)
        wa_template_formula = (
            f'="https://api.whatsapp.com/send?phone=" & D{row_num} & '
            f'"&text=" & ENCODEURL("Shalom / Halo Bapak/Ibu/Saudara/i " & B{row_num} & '
            f'", merupakan suatu kehormatan bagi kami untuk mengundang Anda ke pernikahan kami. Detail undangan digital resmi dapat diakses di: ") & '
            f'ENCODEURL(C{row_num})'
        )
        ws.cell(row=row_num, column=5, value=wa_template_formula).alignment = Alignment(horizontal="left")
        
        # Berikan border dan format font untuk seluruh sel data
        for col_idx in range(1, 6):
            c = ws.cell(row=row_num, column=col_idx)
            c.border = data_border
            if col_idx in [1, 4]:
                c.font = Font(name="Lato", size=10, bold=True)
            elif col_idx in [3, 5]:
                c.font = Font(name="Lato", size=9, color="1F51FF", underline="single") # Hyperlink style
            else:
                c.font = Font(name="Lato", size=10)
    
    # 4. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Cari panjang string terpanjang di tiap kolom (mulai baris 4 ke bawah)
        for cell in col[3:]:
            if cell.value:
                val_str = str(cell.value)
                # Jika formula, taksir panjang link hasil formula untuk estetika lebar kolom
                if val_str.startswith('='):
                    val_str = "https://api.whatsapp.com/send?phone=628123456789&text=Shalom%20Bapak%20Budi..."
                max_len = max(max_len, len(val_str))
        
        # Set lebar kolom dengan sedikit padding keamanan
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    # Simpan file dengan nama baru untuk menghindari file lock error
    output_filename = "Daftar_Undangan_Whatsapp_Formula.xlsx"
    wb.save(output_filename)
    print(f"\n[SUKSES] File Excel Formula berhasil dibuat: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    create_excel_template()
